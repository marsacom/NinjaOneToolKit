#!/user/bin/env python3

# NinjaOneToolKit
# Perform a variety of functions related to devices in NinjaOne/Domain
# Gather values from XLSX Sheet and compare to NinjaOne/Domain
# Update values in XLSX Sheet to reflect devices in/not in NinjaOne/Domain
# List devices in and not in NinjaOne/Domain
# Add devices into NinjaOne/Domain
# Brayden Kukla - 2024

# Generic libraries
import os
import sys
import csv
import warnings
import requests
import subprocess
import argparse
from tabulate import tabulate
from dotenv import load_dotenv
from datetime import datetime
# OpenpyXL libraries
import openpyxl as xl
import openpyxl.styles as xlstyle
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter as xlgcl


endpoint = "https://app.ninjarmm.com/v2/"
oauth_url = "https://app.ninjarmm.com/ws/oauth/token"

# Be sure to change the path in .env 
load_dotenv() 

# Arguements
parser = argparse.ArgumentParser(description="A CLI Application for IT Technicians to access and manage Organizations & Devices in NinjaOne(NinjaRMM) straight from the Command Line", 
                                 prog="NinjaOneToolKit", usage="python3 main.py [options]")
# parser.add_argument('-f', '--file', type=str, help="The XLSX file you wish to compare results from Ninja & AD to...", required=False)
# parser.add_argument('-gf', '--generate-file', help="Use this flag if you wish to generate a XLSX file of results...", action='store_true')

args = parser.parse_args()

# This is the ID of the organization in NinjaOne that your account running the scripts domain  
domain_org_id = os.getenv('DOMAIN_ORG_ID')

# This is specifically to ignore the random warning that is generated when accessing the worksheet via openpyxl (does not affect the script)
warnings.simplefilter('ignore')

user_sel = ''


# Call api endpoint for bearer token, currently this is just uses a machine-to-machine application using client credentials
def get_token():
    data = {
        "grant_type": "client_credentials",
        "client_id": str(os.getenv('CLIENT_ID')),
        "client_secret": str(os.getenv('CLIENT_SECRET')),
        "scope": "monitoring"
    }

    headers = {"Content-Type": "application/x-www-form-urlencoded"}

    token = requests.post(oauth_url, data, headers).json()

    global api_token
    api_token = token["access_token"]


# Get organizations assocaited in NinjaOne
def get_orgs(token):
    global orgs
    global orgs_id
    global user_sel
    
    org_url = endpoint + "organizations/"

    headers = {
        "Accept": "application/json",
        "Authorization": "Bearer " + token,
    }

    organizations = requests.get(org_url, headers=headers).json()

    orgs = []
    orgs_id = []

    print('-'*80 + "\nOrganizations\n" + '-'*80 + '\n')

    c = 0
    for org in organizations:
        c = c + 1
        print(str(c) + ". " + str(org["name"]))
        orgs.append(org["name"])
        orgs_id.append(org["id"])

    print('\n')

    sel =  input("Please select an organization " + "(1-" + str(len(organizations)) + ")... ")
    user_sel = orgs_id[int(sel)-1] # We want the user_sel to be the organization id that they selected, since the ids may not be in numerical order
    print("ORGANIZATION ID : ", str(user_sel))
    get_devices_detailed(token)


# Get detailed information on devices
def get_devices_detailed(token):
    data = [] # Array to store values for displaying in tabulate table
    header = ["System Name", "ID", "Status", "OS", "Brand", "Model", "Serial Number", "Memory", "Processor", "Last Login", "Last Boot Time"] #Headers for tabulate table columns

    headers = {
        "Accept": "application/json",
        "Authorization": "Bearer " + token,
    }

    # Using the built in device filter param to only get detailed info for devices in a specific org
    device_url = endpoint + "devices-detailed/" + "?df=org=" + (str(user_sel) if str(user_sel) != "" else domain_org_id)
    devices = requests.get(device_url, headers=headers).json()

    global ninja_ids
    global ninja_system_names 
    global ninja_status
    global ninja_os_names
    global ninja_system_brands
    global ninja_system_models
    global ninja_system_serials
    global ninja_system_memory
    global ninja_processors
    global ninja_last_login
    global ninja_last_boot

    ninja_ids = []
    ninja_system_names = []
    ninja_status = []
    ninja_os_names = []
    ninja_system_brands = []
    ninja_system_models = []
    ninja_system_serials = []
    ninja_system_memory = []
    ninja_processors = []
    ninja_last_login = []
    ninja_last_boot = []

    print('\n' + '-'*80 + "\nDevices in NinjaOne...\n" + '-'*80)

    if len(devices) >= 1:
        for k in devices:
            try:
                dev_id = (int(k["id"])) if "id" in k else ninja_ids.append(0)
                name = (str(k["systemName"])) if "systemName" in k else ninja_system_names.append("N/A")
                status = (str(k["offline"])) if "offline" in k else ninja_status.append("N/A")
                os = (str(k["os"]["name"])) if "os" in k else ninja_os_names.append("N/A")
                manufacturer = (str(k["system"]["manufacturer"])) if "system" in k and "manufacturer" in k["system"] else ninja_system_names.append("N/A")
                model = (str(k["system"]["model"])) if "system" in k and "model" in k["system"] else ninja_system_models.append("N/A")
                serial = (str(k["system"]["serialNumber"])) if "system" in k and "serialNumber" in k["system"] else ninja_system_serials.append("N/A")
                memory = (round((int(k["memory"]["capacity"]))/(1024 ** 3))) if "memory" in k and "capcity" in k["memory"] else ninja_system_memory.append("N/A")
                processor = (str(k["processors"][0]["name"])) if "processors" in k and "name" in k["processors"][0] else ninja_processors.append("N/A")
                last_login = (str(k["lastLoggedInUser"])) if "lastLoggedInUser" in k else ninja_last_login.append("N/A")
                last_boot = (datetime.fromtimestamp(int(k["os"]["lastBootTime"])).strftime('%m-%d-%Y %H:%M:%S')) if "os" in k and "lastBootTime" in k["os"] else ninja_last_boot.append("N/A")

                ninja_ids.append(dev_id)
                ninja_system_names.append(name)
                ninja_status.append(status)
                ninja_os_names.append(os)
                ninja_system_brands.append(manufacturer)
                ninja_system_models.append(model)
                ninja_system_serials.append(serial)
                ninja_system_memory.append(memory)
                ninja_processors.append(processor)
                ninja_last_login.append(last_login)
                ninja_last_boot.append(last_boot)

                data.append([name, dev_id, status, os, manufacturer, model, serial, memory, processor, last_login, last_boot])
            except Exception as Error:
                print("ERROR: ", Error)

        print(tabulate(data, headers=header, tablefmt='simple_grid'))
    else:
        print("\nThere are no devices currently associated with this organization...\n")
        sys.exit()


# Load excel sheet and gather device info
def get_excel_data():
    global wb
    global ws
    global path
    global xl_ids
    global xl_system_names 
    global xl_row_num
    global xl_ninja_statuses
    global xl_domain_statuses

    # Initialize workbook
    path = os.getenv('XL_PATH')
    wb = xl.load_workbook(path)
    ws = wb[str(os.getenv('XL_WORKSHEET_NAME'))]

    xl_ids = []
    xl_system_names = []
    xl_row_num = []
    xl_ninja_statuses = []
    xl_domain_statuses = []

    l = 1
    # Iterate through the sheet to save values
    for row in ws.iter_rows(min_row=int(os.getenv('XL_MIN_ROW')), max_row=int(os.getenv('XL_MAX_ROW')), values_only=True):
        if row[0] == None:
            pass
        else:
            l = l + 1
            xl_ids.append(row[int(os.getenv('XL_ID_COL'))])
            xl_system_names.append(row[int(os.getenv('XL_SYS_NAME_COL'))])
            xl_row_num.append(l)
            xl_ninja_statuses.append(row[int(os.getenv('XL_NINJA_STATUS_COL'))])
            xl_domain_statuses.append(row[int(os.getenv('XL_DOMAIN_STATUS_COL'))])


# Compare results of devices in NinjaOne to the Excel File and update values in the "Computers" sheet
def compare_res():
    data = [] #Array to store values for displaying in tabulate table
    header = ["Device","In Domain?", "In Ninja?"] #Headers for tabulate table columns    
    ninja_missing = []
    ad_missing = []
    both = []

    print('\n' + '-'*80 + "\nDevices In The Excel File And Their Statuses In NinjaOne & Domain...\n" + '-'*80)

    for i in range(len(xl_system_names)):
        if in_domain(xl_system_names[i]) == False:
            ad_missing.append(xl_system_names[i])
            ws[str(os.getenv('XL_DOMAIN_STATUS_COL_LETTER'))+str(xl_row_num[i])] = 'N'
            if in_ninja(xl_system_names[i]) == False :
                ninja_missing.append((xl_system_names[i]))
                ws[str(os.getenv('XL_NINJA_STATUS_COL_LETTER'))+str(xl_row_num[i])] = 'N'
            else:
                ws[str(os.getenv('XL_NINJA_STATUS_COL_LETTER'))+str(xl_row_num[i])] = 'Y'
        else:
            ws[str(os.getenv('XL_DOMAIN_STATUS_COL_LETTER'))+str(xl_row_num[i])] = 'Y'
            if in_ninja(xl_system_names[i]) == False:
                ninja_missing.append(xl_system_names[i])
                ws[str(os.getenv('XL_NINJA_STATUS_COL_LETTER'))+str(xl_row_num[i])] = 'N'
            else:
                ws[str(os.getenv('XL_NINJA_STATUS_COL_LETTER'))+str(xl_row_num[i])] = 'Y'

        data.append([xl_system_names[i], 'YES' if in_domain(xl_system_names[i]) else 'NO', 'YES' if in_ninja(xl_system_names[i]) else 'NO'])
    
    print(tabulate(data, headers=header, tablefmt='double_grid'))

    # Which devices are missing from NinjaOne & Domain
    for d in range(len(ninja_missing)):
        if ninja_missing[d] in ad_missing:
            both.append(ninja_missing[d])

    #Write to the log file and save changes made to the workbook
    write_to_file(ninja_missing, ad_missing, both)
    wb.save(path)


# Generate an excel file with all devices in a specified organization and their information
def generate_xlsx():

    print("Generating XLSX file with results of devices in Ninja...")

    # Initizlize & create workbook/worksheet
    wb = xl.Workbook()
    ws = wb.active

    # First lets add our headers to the XLSX file
    ws['A1'] = "System Name"
    ws['B1'] = "Status"
    ws['C1'] = "OS"
    ws['D1'] = "Brand"
    ws['E1'] = "Model"
    ws['F1'] = "Serial #"
    ws['G1'] = "Memory (GB)"
    ws['H1'] = "Processor"
    ws['I1'] = "Last Login"
    ws['J1'] = "Last Boot Time"

    # Iterate through the sheet to save values
    row = 1
    try: 
        for i in range(0,len(ninja_system_names)) :
            row = row + 1
            ws["A" + str(row)] = ninja_system_names[i]
            ws["B" + str(row)] = "Offline" if ninja_status[i] == "True" else "Online"
            ws["C" + str(row)] = ninja_os_names[i]
            ws["D" + str(row)] = ninja_system_brands[i]
            ws["E" + str(row)] = ninja_system_models[i]
            ws["F" + str(row)] = ninja_system_serials[i]
            ws["G" + str(row)] = ninja_system_memory[i]
            ws["H" + str(row)] = ninja_processors[i]
            ws["I" + str(row)] = ninja_last_login[i]
            ws["J" + str(row)] = ninja_last_boot[i]
        
        # Now lets attempt to format the sheet
        try: 
            # Create and style table
            table = Table(displayName='NinjaDevices', ref='A1:' + xlgcl(ws.max_column) + str(ws.max_row))
            table.tableStyleInfo = TableStyleInfo(name="TableStyle", showFirstColumn=True, showLastColumn=True,
                                showRowStripes=True, showColumnStripes=True)
            ws.add_table(table)  

            for row in ws.iter_rows(): # Lets set the allignment properties
                for cell in row:
                    cell.alignment = xlstyle.Alignment(vertical='center', horizontal='center')

            for column in ws.columns: # Now lets set the column widths
                length = 0
                letter =  column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > length:
                            length = len(cell.value)
                    except:
                        pass
                width = (length + 2) # Add a little extra room in the cell to work with
                ws.column_dimensions[letter].width = width
        except Exception as Error:
            print("\nERROR: Unable to format XLSX file - ", Error)  

        # Concat the selected org name with the file name, user_sel is the selected orgs id, run a function to get the org name from its ID
        org_name = orgs[orgs_id.index(user_sel)]
        wb_name = org_name.replace(" ", "-") + "-Ninja-Devices-" + str(datetime.now().strftime("%Y-%m-%d %H:%M")) + ".xlsx"
        
        folder_path = os.getcwd() + "\\XLSX Results\\" 
        full_path =  folder_path + wb_name

        if not os.path.exists(folder_path): # Lets check if the folder already exists first
            os.mkdir(os.getcwd() + "\\XLSX Results")

        wb.save(full_path) # Finally lets save the workbook

        print("\nSuccessfully generated XLSX file, file can be found at..." + full_path + "\n")
    except Exception as Error:
        print("ERROR: Unable to generate XLSX file - ", Error)


# Get all computers associated with Active Directory
def get_ad_computers():
    # First gather all devices via Get-ADComp cmdlet then parse the information from the CSV
    cmd = " Get-ADComputer -Filter * -Properties IPv4Address | Export-Csv " + os.getcwd() + "\\computers.csv"
    p = subprocess.Popen('powershell -command' + cmd)
    p.communicate()

    # Now let's parse the CSV and check/store values for later comparisons
    data = [] #Array to store values for displaying in tabulate table
    header = ["System Name", "DNS Name", "IP Address"] #Headers for tabulate table columns1
    file = os.getcwd() + "\\computers.csv"

    global ad_rows
    global ad_dns
    global ad_ips
    global ad_names

    ad_rows = []
    ad_dns = []
    ad_ips = []
    ad_names = []

    try:
        with open(file, 'r') as csvfile:
            reader = csv.reader(csvfile)
            
            for row in reader:
                ad_rows.append(row)

            # This just deletes the first 2 items in the list to get rid of the bullshit info we dont want
            for i in range(2):
                ad_rows.pop(0)

            print('\n' + '-'*80 + "\nDevices in the Domain...\n" + '-'*80)

            for row in ad_rows:
                ad_names.append(row[4])
                ad_dns.append(row[1])
                if row[3] == '': # Some of the IPs are unknown in the domain for some reason, this is just to check 
                    ad_ips.append('   UNKNOWN  ')
                else:
                    ad_ips.append(row[3])
                data.append([row[4], row[1], 'UNKNOWN' if row[3] == '' else row[3]])

            print(tabulate(data, headers=header, tablefmt="double_grid"))    
    except FileNotFoundError:
        print("ERROR: CSV file not found...")


# WIP : Add device to NinjaOne Organization
def add_to_ninja():
    if os.path.exists(str(os.getenv('INSTALL_PATH'))):
        cmd = ""
        p = subprocess.Popen('powershell -command' + cmd)
        p.communicate


#Comparison functions
def in_ninja(device):
    if device in ninja_system_names:
        return True
    else:
        return False


def in_domain(device):
    if device in ad_names:
        return True
    else:
        return False


# Write results to results.txt file in the specified log path
def write_to_file(ninja_missing, ad_missing, both):

    dev_lbl = "Device: "
    folder_path = os.getcwd() + "\\Logs\\"
    file_path = os.path.join(folder_path, "results.txt")

    if not os.path.exists(folder_path): # Lets check if our file already exists...
        os.mkdir(folder_path)
    else:
        pass

    try:    
        with open(file_path, "w") as f:
            f.write("NinjaOne\n" + "------------\n")
            for i in range(len(ninja_missing)):
                f.write(dev_lbl + ninja_missing[i] + " has NOT yet joined NinjaOne... \n")
            f.write("\nDomain\n" + "--------\n")
            for d in range(len(ad_missing)):
                f.write(dev_lbl + ad_missing[d] + " has NOT yet joined the Domain... \n")
            f.write("\nNinjaOne & Domain\n" + "----------------\n")
            for k in range(len(both)):
                f.write(dev_lbl + both[k] + " has NOT yet joined the Domain or NinjaOne... \n")
            f.write("\nSUCCESS: Script completed at - " + str(datetime.now()))

            print('-'*80 + "\n\nSUCCESS: Results have been saved in " + file_path + '...\n')   

    except FileNotFoundError:
        print("\nERROR: File not found. Unable to create log file and/or log folder, please check permissions on your CWD...\n")  


def device_in_ninja_not_domain():
    data = []
    header = ["System Name"]

    for i in range(len(ninja_system_names)):
        if not in_domain(ninja_system_names[i]):
            data.append([ninja_system_names[i]])
        else:
            pass

    print("\nDevices in NinjaOne but NOT the domain...\n")
    print(tabulate(data, headers=header, tablefmt='double_grid'))


def device_in_domain_not_ninja():
    data = []
    header = ["System Name"]
    
    for i in range(len(ad_names)):
        if not in_ninja(ad_names[i]):
            data.append([ad_names[i]])
        else:
            pass
    print("\nDevices in the domain but NOT NinjaOne...\n")
    print(tabulate(data, headers=header, tablefmt='double_grid'))


# Main
def main():
    get_token()
    print("\nStarting NinjaOneToolKit v.1.1...")   
    print('-'*80 + "\n 1: List all devices in Ninja\n", "2: List all devices in the domain\n", "3: List all devices that are in Ninja but NOT the domain\n", 
          "4: List all devices that are in the domain but NOT Ninja\n", "5: List devices in Ninja & the domain and compare with XLSX file\n", 
          "6: Generate XLSX file of devices in Ninja\n", "7: Add computer to NinjaOne (WIP)\n")

    choice = int(input("Please select an option from the list above (1-5)... "))
    
    if choice == 1: # List all devices in NinjaOne
        get_orgs(api_token)
    elif choice == 2: # List all devices in the Domain
        get_ad_computers()
    elif choice == 3: # List all devices that are in Ninja but NOT the domain
        get_ad_computers()
        get_devices_detailed(api_token)
        device_in_ninja_not_domain()
    elif choice == 4: # List all devices that are in the domain but NOT Ninja
        get_ad_computers()
        get_devices_detailed(api_token)
        device_in_domain_not_ninja()    
    elif choice == 5: # Get devices in NinjaOne and the Domain and compare with the XLSX Sheet
        get_ad_computers()
        get_devices_detailed(api_token)  
        get_excel_data()
        compare_res()
    elif choice == 6: # Generate XLSX file from results of devicves in Ninja
        get_orgs(api_token)
        generate_xlsx()
    elif choice == 7: # Add computer to NinjaOne
        print("\nThis feature is currently being developed and is unavailable...")
    else:
        print("\nERROR: Please re-run the script and enter a valid value, 1-5")


if __name__=="__main__":
    main()
else:
    print("ERROR: Unknown error occurred, exiting...\n")
    sys.exit()